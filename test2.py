import os                                                               # модуль по раюоте с директориями
import openpyxl                                                         # модуль по работе с excel
from docxtpl import DocxTemplate                                        # модуль по работе с docx

#Открываем директорию с запущенным скриптов и по этому пути создаем 3 папки для будущих документов

starting_path = os.getcwd()                                             # путь до рабочей пргограммы
print(starting_path)

'''
os.makedirs(path + '/!!!ACT!!!', exist_ok = True)                       # создаем директорию для актов
path_act = path + '/!!!ACT!!!'
os.makedirs(path + '/!!!CONTRACT!!!', exist_ok = True)                  # договоров
path_contract = path + '/!!!CONTRACT!!!'
os.makedirs(path + '/!!!CERTIFICATE!!!', exist_ok = True)               # сертификатов
path_certificate = path + '/!!!CERTIFICATE!!!'

path_date = path + '/!!!DATE!!!'
path_sketches = path + '/SKETCHES_OF_DOCUMENTS'
'''

# Из начальной директории с рабочей программой переходим в директорию с файлом информации, так как в функции вызова EXCEL файла нет возможности открыть файл в иной директори
# Входным является параметр пути до директории
# Выходным параметром является лист с данными

def path_to_excel():

    global starting_path

    path_date = starting_path + '/!!!DATE!!!'
    os.chdir(path_date)
    wb = openpyxl.load_workbook('21.07.xlsx')                       #открытие файла Excel
    sheet = wb['Лист1']
    os.chdir(starting_path)
    
    return sheet

def ACT(i, k):

    global starting_path
    global cols

    FLAG_2 = 1                                      # проверка из какого модуля была вызвана подпрограмма

    new_local_path = local_path(FLAG_2)

    X = [   'NUMBER',                               #0
            'NAME',                                 #1
            'NAME_RUS',                             #2
            'COUNTRY',                              #3
            'COUNTRY_RUS',                          #4
            'DATE_OF_BIRTH',                        #5
            'DEPARTMENT',                           #6
            'DEPARTMENT_RUS',                       #7
            'NUMBER_OF_CONTRACT',                   #8
            'CERTIFICATE_NUMBER',                   #9
            'DATE_OF_ISSUE',                        #10
            'WEEKS',                                #11
            'DATE_FROM',                            #12
            'DATE_TO',                              #13
            'SUM',                                  #14
            'SUM_NAME',                             #15
            'SUM_NAME_RUS'                          #16
        ]
    
    #print(X)

    for j in range(0, cols):
        cell = k.cell(row = i + 1, column = j + 1)
        X[j] = cell.value
    
    print(X)
    
    
    
    #print(X)
    
    
    context =   {       'NUMBER_OF_CONTRACT' :      X[8],
                        'DATE_FROM' :               X[12],
                        'DATE_TO' :                 X[13],
                        'COUNRTY' :                 X[3],
                        'COUNTRY_RUS' :             X[4],
                        'SUM' :                     X[14],
                        'SUM_NAME' :                X[15],
                        'SUM_NAME_RUS' :            X[16],
                        'NAME' :                    X[1],
                        'NAME_RUS' :                X[2]
                }

    doc = open_sketch(FLAG_2, new_local_path)
    doc.render(context)
    doc.save(str(X[0]) + '. ' + X[1] + '.docx')
    os.chdir(starting_path)

def CONTRACT(i, k):

    global starting_path
    global cols

    FLAG_2 = 2                                      # проверка из какого модуля была вызвана подпрограмма

    new_local_path = local_path(FLAG_2)

    X = [   'NUMBER',                               #0
            'NAME',                                 #1
            'NAME_RUS',                             #2
            'COUNTRY',                              #3
            'COUNTRY_RUS',                          #4
            'DATE_OF_BIRTH',                        #5
            'DEPARTMENT',                           #6
            'DEPARTMENT_RUS',                       #7
            'NUMBER_OF_CONTRACT',                   #8
            'CERTIFICATE_NUMBER',                   #9
            'DATE_OF_ISSUE',                        #10
            'WEEKS',                                #11
            'DATE_FROM',                            #12
            'DATE_TO',                              #13
            'SUM',                                  #14
            'SUM_NAME',                             #15
            'SUM_NAME_RUS'                          #16
        ]
    
    #print(X)

    for j in range(0, cols):
        cell = k.cell(row = i + 1, column = j + 1)
        X[j] = cell.value
        if j == 5:
    	    X[j] = str(X[j])
    	    X[j] = X[j][:4]
    
    print(X)
  
    #CHECK(X, k)
 
    #print(X)
    
    
    context =   {       'NUMBER_OF_CONTRACT' :      X[8],
                        'DATE_FROM' :               X[12],
                        'DATE_TO' :                 X[13],
                        'COUNRTY' :                 X[3],
                        'COUNTRY_RUS' :             X[4],
                        'SUM' :                     X[14],
                        'SUM_NAME' :                X[15],
                        'SUM_NAME_RUS' :            X[16],
                        'NAME' :                    X[1],
                        'NAME_RUS' :                X[2],
                        'DATE_OF_BIRTH' :           X[5],
			'DEPARTMENT' :		X[6],
			 'DEPARTMENT_RUS' :                      X[7]

                }

    doc = open_sketch(FLAG_2, new_local_path)
    doc.render(context)
    doc.save(str(X[0]) + '. ' + X[1] + '.docx')
    os.chdir(starting_path)

def CERTIFICATE(i, k):

    global starting_path
    global cols

    FLAG_2 = 3                                      # проверка из какого модуля была вызвана подпрограмма

    new_local_path = local_path(FLAG_2)

    X = [   'NUMBER',                               #0
            'NAME',                                 #1
            'NAME_RUS',                             #2
            'COUNTRY',                              #3
            'COUNTRY_RUS',                          #4
            'DATE_OF_BIRTH',                        #5
            'DEPARTMENT',                           #6
            'DEPARTMENT_RUS',                       #7
            'NUMBER_OF_CONTRACT',                   #8
            'CERTIFICATE_NUMBER',                   #9
            'DATE_OF_ISSUE',                        #10
            'WEEKS',                                #11
            'DATE_FROM',                            #12
            'DATE_TO',                              #13
            'SUM',                                  #14
            'SUM_NAME',                             #15
            'SUM_NAME_RUS'                          #16
        ]
    
    #print(X)

    for j in range(0, cols):
        cell = k.cell(row = i + 1, column = j + 1)
        X[j] = cell.value
        if j == 5:
    	    X[j] = str(X[j])
    	    X[j] = X[j][:10]
    
    print(X)
    
    #CHECK(X, k)
    
    #print(X)
    
    
    context =   {       'NAME' :                    X[1],
                        'NAME_RUS' :                X[2],
                        'DATE_OF_BIRTH' :           X[5],
                        'DEPARTMENT' :              X[6],
                        'DEPARTMENT_RUS' :          X[7],
                        'CERTIFICATE_NUMBER' :      X[9]

                }

    doc = open_sketch(FLAG_2, new_local_path)
    doc.render(context)
    doc.save(str(X[0]) + '. ' + X[1] + '.docx')
    os.chdir(starting_path)

def CHECK(X, k):

    if X[11] == 'FIRST 2 WEEKS':
        N = 3
    if X[11] == 'SECOND 2 WEEKS':
        N = 7
    if X[11] == '3 WEEKS':
        N = 11
    print(N)
    for i in range(14, 19):
        cell = k.cell(row = N, column = i)
        X[i - 2] = cell.value

def open_sketch(FLAG_2, new_local_path):
    
    global starting_path

    path_sketch = starting_path + '/!!!SKETCHES_OF_DOCUMENTS!!!'
    os.chdir(path_sketch)

    if FLAG_2 == 1:
        doc = DocxTemplate('Akt.docx')                                #открытие фалйа Word
    if FLAG_2 == 2:
        doc = DocxTemplate('Dogovor.docx')
    if FLAG_2 == 3:
        doc = DocxTemplate('sketch_of_certificate.docx')

    os.chdir(new_local_path)

    return doc

def local_path(FLAG_2):

    global starting_path

    if FLAG_2 == 1:
        direction = '/!!!ACT!!!'
    if FLAG_2 == 2:
        direction = '/!!!CONTRACT!!!'
    if FLAG_2 == 3:
        direction = '/!!!CERTIFICATE!!!'

    new_local_path = starting_path + direction
    os.makedirs(new_local_path, exist_ok = True)                       # создаем директорию для актов
    os.chdir(new_local_path)

    return new_local_path

#**********************************************************************************************#
#                                                                                              #
"""   !!! ТЕЛО ПРОГРАММЫ !!!   """
#                                                                                              #
#**********************************************************************************************#

k = path_to_excel()                                                             # функция с таблицой данных
rows = k.max_row
print(rows)
cols = 12

#**********************************************************************************************#
#                                                                                              #
"""
FLAG отвечает за выбор нужной функции:
1 - только акты
2 - договоры
3 - сертификаты
4 - все
"""
#                                                                                              #
#**********************************************************************************************#

FLAG_1 = 3

FIRST_ROW = 2

#TEST_NUMBER_OF_ROWS = 1
"""FIRST_ROW + rows - 2"""

for i in range(FIRST_ROW - 1, FIRST_ROW + rows - 2):

    if FLAG_1 == 1:
        ACT(i ,k)

    if FLAG_1 == 2:
        CONTRACT(i, k)

    if FLAG_1 == 3:
        CERTIFICATE(i, k)

    if FLAG_1 == 4:
        ACT(i ,k)
        CONTRACT(i ,k)
        CERTIFICATE(i ,k)